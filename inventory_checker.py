import os
import pandas as pd
import pdfplumber
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import re
import argparse
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import get_close_matches
import requests

class InventoryChecker:
    def __init__(self):
        self.base_url = "https://southernstarz.com"
        self.driver = None
        self.max_retries = 3
        self.page_load_timeout = 30
        self.wait = None
        self.progress_callback = None  # Callback for progress updates
        
        # Define SKUs that are allowed to share URLs (intentionally)
        self.allowed_duplicate_skus = {
            # NEWBLOOD Chardonnay
            'S1NBCHANAC': ['S1NBCHANAE', 'S1NBCHANAF', 'S1NBCHANAG'],
            'S1NBCHANAE': ['S1NBCHANAC', 'S1NBCHANAF', 'S1NBCHANAG'],
            'S1NBCHANAF': ['S1NBCHANAC', 'S1NBCHANAE', 'S1NBCHANAG'],
            'S1NBCHANAG': ['S1NBCHANAC', 'S1NBCHANAE', 'S1NBCHANAF'],
            
            # NEWBLOOD Red Blend
            'S1NBRBLNAE': ['S1NBRBLNAF', 'S1NBRBLNAG'],
            'S1NBRBLNAF': ['S1NBRBLNAE', 'S1NBRBLNAG'],
            'S1NBRBLNAG': ['S1NBRBLNAE', 'S1NBRBLNAF'],
            
            # NEWBLOOD Rose
            'S1NBROSNAF': ['S1NBROSNAG'],
            'S1NBROSNAG': ['S1NBROSNAF'],
            
            # Antucura SV Pukara Cabernet Sauvignon 2015
            'S1ANTCSS15': ['S1ANTMAS15'],
            'S1ANTMAS15': ['S1ANTCSS15'],
            
            # Casas del Bosque Collection Chardonnay 2024
            'S1CDBCH24': ['S1CDBCHC24'],
            'S1CDBCHC24': ['S1CDBCH24'],
            
            # Greenock Creek Alice's Shiraz 2020
            'S1GCALSH20': ['S1GCSHI20'],
            'S1GCSHI20': ['S1GCALSH20'],
            
            # David Finlayson GS Cabernet Sauvignon 2020
            'S1EDGGCS20-6': ['S1EDGGCSM20'],
            'S1EDGGCSM20': ['S1EDGGCS20-6']
        }
        
        # Common varietal terms for matching
        self.varietal_terms = {
            'chardonnay': 'chardonnay',
            'chard': 'chardonnay',
            'cab': 'cabernet',
            'cabernet': 'cabernet',
            'cabernet sauvignon': 'cabernet',
            'merlot': 'merlot',
            'malbec': 'malbec',
            'pinot': 'pinot noir',
            'pinot noir': 'pinot noir',
            'pn': 'pinot noir',
            'sauvignon': 'sauvignon blanc',
            'sauvignon blanc': 'sauvignon blanc',
            'sauv blanc': 'sauvignon blanc',
            'sb': 'sauvignon blanc',
            'shiraz': 'shiraz',
            'syrah': 'shiraz',
            'grenache': 'grenache',
            'pinotage': 'pinotage',
            'riesling': 'riesling',
            'chenin': 'chenin blanc',
            'chenin blanc': 'chenin blanc',
            'carmenere': 'carmenere',
            'cab franc': 'cabernet franc',
            'cabernet franc': 'cabernet franc',
            'rose': 'rose',
            'rosé': 'rose',
            'sparkling': 'sparkling',
            'red blend': 'red blend',
            'muscat': 'muscat',
            'tokay': 'tokay',
            'tawny': 'tawny',
            'roussanne': 'roussanne'
        }
        
        # Add direct SKU to URL mappings
        self.sku_url_mapping = {
            'S1ALMMA2275': 'https://southernstarz.com/wines/almarada-malbec-2022/',
            'S1ANTCHNV': 'https://southernstarz.com/wines/antucura-cherie-sparkling-nv/',
            'S1NBCHANAC': 'https://southernstarz.com/wines/newblood-chardonnay/',
            'S1NBCHANAE': 'https://southernstarz.com/wines/newblood-chardonnay/',
            'S1NBCHANAF': 'https://southernstarz.com/wines/newblood-chardonnay/',
            'S1NBRBLNAE': 'https://southernstarz.com/wines/newblood-red-blend/',
            'S1NBRBLNAF': 'https://southernstarz.com/wines/newblood-red-blend/',
            'S1NBROSNAF': 'https://southernstarz.com/wines/newblood-rose/',
            'S1ANTCSS15': 'https://southernstarz.com/wines/antucura-sv-pukara-cabernet-sauvignon-2015/',
            'S1ANTMA22': 'https://southernstarz.com/wines/antucura-barrandica-malbec-2022/',
            'S1CDBCAR22': 'https://southernstarz.com/wines/casas-del-bosque-reserva-carmenere-2022/',
            'S1CDBCH24': 'https://southernstarz.com/wines/casas-del-bosque-collection-chardonnay-2024/',
            'S1CDBSBC24': 'https://southernstarz.com/wines/casas-del-bosque-single-vineyard-collection-sauvignon-blanc-2024/',
            
            # New mappings with blank URLs (to be filled in manually)
            'S1ANTCA20': 'https://southernstarz.com/wines/antucura-calcura-2020/',
            'S1ANTMAS15': 'https://southernstarz.com/wines/antucura-sv-pukara-cabernet-sauvignon-2015/',
            'S1BLACB23': 'https://southernstarz.com/wines/black-pearl-chenin-blanc-2023/',
            'S1BLACB24': 'https://southernstarz.com/wines/black-pearl-chenin-blanc-2024/',
            'S1CDBCA23': 'https://southernstarz.com/wines/casas-del-bosque-collection-carmenere-2023/',
            'S1CDBCHC24': 'https://southernstarz.com/wines/casas-del-bosque-collection-chardonnay-2024/',
            'S1CDBCHR23': 'https://southernstarz.com/wines/casas-del-bosque-reserva-chardonnay-2023/',
            'S1CDBGCH18': 'https://southernstarz.com/wines/casas-del-bosque-gran-reserva-chardonnay-2018/',
            'S1CDBGCS23': 'https://southernstarz.com/wines/casas-del-bosque-gran-reserva-cabernet-sauvignon-2023/',
            'S1CDBGPN23': 'https://southernstarz.com/wines/casas-del-bosque-gran-reserva-pinot-noir-2023/',
            'S1CDBGSY21': 'https://southernstarz.com/wines/casas-del-bosque-gran-reserva-syrah-2021/',
            'S1CDBSBL24': 'https://southernstarz.com/wines/casas-del-bosque-la-cantera-sauvignon-blanc-2024/',
            'S1EDGCCF22': 'https://southernstarz.com/wines/david-finlayson-camino-africana-cabernet-franc-2022/',
            'S1EDGCCH21': 'https://southernstarz.com/wines/david-finlayson-camino-africana-chardonnay-2021/',
            'S1EDGCPN18': 'https://southernstarz.com/wines/david-finlayson-camino-africana-pinot-2018/',
            'S1EDGDCEA21': 'https://southernstarz.com/wines/david-finlayson-cab-et-al-2021/',
            'S1EDGDCH23': 'https://southernstarz.com/wines/david-finlayson-chardonnay-2023/',
            'S1EDGDCS21': 'https://southernstarz.com/wines/david-finlayson-cabernet-sauvignon-2021/',
            'S1EDGGCS20-6': 'https://southernstarz.com/wines/david-finlayson-gs-cabernet-sauvignon-2020/',
            'S1EDGGCS21-6': 'https://southernstarz.com/wines/david-finlayson-gs-cabernet-sauvignon-2021/',
            'S1EDGPP22': 'https://southernstarz.com/wines/david-finlayson-the-pepper-pot-2022/',
            'S1DFME19': 'https://southernstarz.com/wines/downes-family-merlot-2019/',
            'S1DFME20': 'https://southernstarz.com/wines/downes-family-merlot-2020/',
            'S1DFSB22': 'https://southernstarz.com/wines/downes-sanctuary-peak-sauvignon-blanc-2022/',
            'S1DOWMBME20': 'https://southernstarz.com/wines/downes-mt-bullet-merlot-2020/',
            'S1EASSB23': 'https://southernstarz.com/wines/earthsong-sauvignon-blanc-2023/',
            'S1EASSB24': 'https://southernstarz.com/wines/earthsong-sauvignon-blanc-2024/',
            'S1GCALSH20': 'https://southernstarz.com/wines/greenock-creek-alices-shiraz-2020/',
            'S1GCALSH21': 'https://southernstarz.com/wines/greenock-creek-alices-shiraz-2021/',
            'S1GCJSH19': 'https://southernstarz.com/wines/greenock-creek-jaensch-shiraz-2019/',
            'S1GCJSH20': 'https://southernstarz.com/wines/greenock-creek-jaensch-shiraz-2020/',
            'S1GCRSH16': 'https://southernstarz.com/wines/greenock-creek-roennfeldt-shiraz-2016/',
            'S1GCSASH20': 'https://southernstarz.com/wines/greenock-creek-seven-acre-shiraz-2020/',
            'S1GCSHI20': 'https://southernstarz.com/wines/greenock-creek-alices-shiraz-2020/',
            'S1GCSSH19': 'https://southernstarz.com/wines/greenock-creek-stone-block-shiraz-2019/',
            'S1MIHASB24': 'https://southernstarz.com/wines/miha-sauvignon-blanc-2024/',
            'S1NBCHANAG': 'https://southernstarz.com/wines/newblood-chardonnay/',
            'S1NBRBLNAG': 'https://southernstarz.com/wines/newblood-red-blend/',
            'S1NGCHD21': 'https://southernstarz.com/wines/nugan-estate-sv-drovers-hut-chardonnay-2021/',
            'S1NGCHER23': 'https://southernstarz.com/wines/nugan-estate-third-generation-chardonnay-2023/',
            'S1NGCHER24': 'https://southernstarz.com/wines/nugan-estate-third-generation-chardonnay-2024/',
            'S1NGCSST22': 'https://southernstarz.com/wines/nugan-estate-sv-stompers-cabernet-sauvignon-2022/',
            'S1NGCSST23': 'https://southernstarz.com/wines/nugan-estate-sv-stompers-cabernet-sauvignon-2023/',
            'S1NGKVCH19': 'https://southernstarz.com/wines/nugan-estate-kv-frascas-lane-vineyard-chardonnay-2019/',
            'S1NGSHR-21': 'https://southernstarz.com/wines/nugan-estate-third-generation-shiraz-2021/',
            'S1NGSHSC21': 'https://southernstarz.com/wines/nugan-estate-sv-scruffys-shiraz-2021/',
            'S1NUADSH14': 'https://southernstarz.com/wines/nugan-estate-alfredo-dried-grape-shiraz-2014/',
            'S1OLIRSH20': 'https://southernstarz.com/wines/oliverhill-red-silk-shiraz-2020/',
            'S1REDCHB24': 'https://southernstarz.com/wines/painted-wolf-the-den-chenin-blanc-2024/',
            'S1REDCS21': 'https://southernstarz.com/wines/painted-wolf-the-den-cabernet-sauvignon-2021/',
            'S1REDGPN21': 'https://southernstarz.com/wines/painted-wolf-guillermo-pinotage-2021/',
            'S1REDPNN22': 'https://southernstarz.com/wines/painted-wolf-the-den-pinotage-2022-2/',
            'S1REDPNN23': 'https://southernstarz.com/wines/painted-wolf-the-den-pinotage-2023/',
            'S1REDPR20': 'https://southernstarz.com/wines/painted-wolf-ros-pinotage-rose-2020/',
            'S1RFRI1021': 'https://southernstarz.com/wines/rieslingfreak-no-10-eden-clare-valley-riesling-2021/',
            'S1RFRI1023': 'https://southernstarz.com/wines/rieslingfreak-no-10-eden-clare-valley-riesling-2023/',
            'S1RFRI1222': 'https://southernstarz.com/wines/rieslingfreak-no-12-flaxman-eden-valley-2022/',
            'S1RFRI223': 'https://southernstarz.com/wines/rieslingfreak-no-2-polish-hill-river-riesling-2023/',
            'S1RFRI3323': 'https://southernstarz.com/wines/rieslingfreak-no-33-clare-valley-riesling-2023/',
            'S1SHEPN20': 'https://southernstarz.com/wines/sherwood-pinot-noir-2020/',
            'S1TAIBAL21': 'https://southernstarz.com/wines/tait-the-ball-buster-2021/',
            'S1TAIBSH22': 'https://southernstarz.com/wines/tait-border-crossing-shiraz-2022/',
            'S1TAISH18': 'https://southernstarz.com/wines/tait-basket-pressed-shiraz-2018/',
            'S1TAIWIR19': 'https://southernstarz.com/wines/tait-the-wild-ride-2019/',
            'S1TDCPSH22': 'https://southernstarz.com/wines/thistledown-the-cunning-plan-shiraz-2022/',
            'S1TDCPSH23': 'https://southernstarz.com/wines/thistledown-the-cunning-plan-shiraz-2023/',
            'S1TDFHGR22': 'https://southernstarz.com/wines/thistledown-fool-on-the-hill-grenache-2022/',
            'S1TDGOGR22': 'https://southernstarz.com/wines/thistledown-gorgeous-grenache-2022/',
            'S1TDGOGRB23': 'https://southernstarz.com/wines/thistledown-gorgeous-grenache-blanc-2023/',
            'S1TDGOGRB24': 'https://southernstarz.com/wines/thistledown-gorgeous-grenache-blanc-2024/',
            'S1TDGOSH22': 'https://southernstarz.com/wines/thistledown-gorgeous-shiraz-2022/',
            'S1TDQUSH21': 'https://southernstarz.com/wines/thistledown-the-quickening-shiraz-2021/',
            'S1TDSEGR23': 'https://southernstarz.com/wines/thistledown-shes-electric-grenache-2023/',
            'S1TDSTGR22': 'https://southernstarz.com/wines/thistledown-sands-of-time-old-vine-grenache-2022/',
            'S1TDTDGR23': 'https://southernstarz.com/wines/thistledown-the-thorny-devil-grenache-2023/',
            'S1TDVGR22': 'https://southernstarz.com/wines/thistledown-the-vagabond-grenache-2022/',
            'S1TDVGR23': 'https://southernstarz.com/wines/thistledown-the-vagabond-grenache-2023/',
            'S1TDWED22': 'https://southernstarz.com/wines/thistledown-where-eagles-dare-shiraz-2022/',
            'S1TDWWKRW22': 'https://southernstarz.com/wines/thistledown-walking-with-kings-roussanne-2022/',
            'S1MTFSB23': 'https://southernstarz.com/wines/mount-fishtail-sur-lie-sauvignon-blanc-2023/',
            'S1MTFSB24': 'https://southernstarz.com/wines/mount-fishtail-sur-lie-sauvignon-blanc-2024/',
            'S1RLBMUNV': 'https://southernstarz.com/wines/rl-buller-fine-muscat-nv-375-ml/',
            'S1RLBTA750': 'https://southernstarz.com/wines/rl-buller-victoria-tawny-nv-750ml/',
            'S1RLBTONV': 'https://southernstarz.com/wines/rl-buller-fine-tokay-nv-375-ml/',
            'S1VABRO19': 'https://southernstarz.com/wines/vina-alicia-brote-negro-malbec-2019/',
            'S1VACMA20': 'https://southernstarz.com/wines/vina-alicia-las-compuertas-malbec-2020/',
            'S1VAMOR17': 'https://southernstarz.com/wines/vina-alicia-morena-2017/',
            'S1VAMOR18': 'https://southernstarz.com/wines/vina-alicia-morena-cabernet-sauvignon-2018/',
            'S1VAMOR19': 'https://southernstarz.com/wines/vina-alicia-morena-cabernet-sauvignon-2019/',
            'S1VAPCS13': 'https://southernstarz.com/wines/vina-alicia-paso-de-piedra-cabernet-sauvignon-2013/',
            'S1VAPCS21': 'https://southernstarz.com/wines/vina-alicia-paso-de-piedra-cabernet-sauvignon-2021/',
            'S1VAPMA20': 'https://southernstarz.com/wines/vina-alicia-paso-de-piedra-malbec-2020/',
            'S1VAPMA21': 'https://southernstarz.com/wines/vina-alicia-paso-de-piedra-malbec-2021/',
            'S1VATIA19': 'https://southernstarz.com/wines/vina-alicia-tiara-2019/',
            'S1VATIA21': 'https://southernstarz.com/wines/vina-alicia-tiara-2021/',
            'S1WAWTE20': 'https://southernstarz.com/wines/water-wheel-estate-2020/',
            'S1WILRCH16': 'https://southernstarz.com/wines/wildberry-reserve-chardonnay-2016/',
            'S1CDBPNC24': 'https://southernstarz.com/wines/casas-del-bosque-collection-single-vineyard-pinot-noir-2024/',
            'S1CDBCSC23': 'https://southernstarz.com/wines/casas-del-bosque-collection-cabernet-sauvignon-2023/',
            'S1NGCAB23': 'https://southernstarz.com/wines/nugan-estate-third-generation-cabernet-sauvignon-2023/',
            'S1ALMCS23': 'https://southernstarz.com/wines/almarada-cabernet-sauvignon-2023/',
            'S1RLBTONV': 'https://southernstarz.com/wines/rl-buller-fine-tokay-nv-375-ml/',
            'S1NGACS19': 'https://southernstarz.com/wines/nugan-alcira-cabernet-sauvignon-2019/',
            'S1EDGDCS22': 'https://southernstarz.com/wines/david-finlayson-cabernet-sauvignon-2022/',
            'S1EDGGCSM20': 'https://southernstarz.com/wines/david-finlayson-gs-cabernet-sauvignon-2020/',
            'S1TDWWKRW23': 'https://southernstarz.com/wines/thistledown-walking-with-kings-roussanne-grenache-blanc-2023/',
            'S1TDSTGR23': 'https://southernstarz.com/wines/thistledown-sands-of-time-old-vine-grenache-2023/',
            'S1RLBTA750': 'https://southernstarz.com/wines/rl-buller-victoria-tawny-nv-750ml/',
            'S1RLBMUNV': 'https://southernstarz.com/wines/rl-buller-fine-muscat-nv-375-ml/',
            'S1TAIWIR21': 'https://southernstarz.com/wines/tait-the-wild-ride-2021/',
            'S1TDCMGR23': 'https://southernstarz.com/wines/thistledown-this-charming-man-grenache-2023/',
            'S1BLAORM22': 'https://southernstarz.com/wines/black-pearl-the-mischief-maker-2022/',
            'S1ANTCF23': 'https://southernstarz.com/wines/antucura-barrandica-cabernet-franc-2023/',
            'S1MTFSB23': 'https://southernstarz.com/wines/mount-fishtail-sur-lie-sauvignon-blanc-2023/',
            'S1MTFSB24': 'https://southernstarz.com/wines/mount-fishtail-sur-lie-sauvignon-blanc-2024/',

        }
        
        self.producers = [
            'Mount Fishtail',
            'Sherwood',
            'Stratum',
            'Black Pearl Vineyards',
            'Downes Family Vineyards',
            'David Finlayson',
            'Luddite',
            'Painted Wolf',
            'Thistledown',
            'Nugan Estate',
            'RL Buller',
            'Water Wheel',
            'Casas del Bosque',
            'Earthsong',
            'Miha',
            'Almarada',
            'Antucura',
            'Vina Alicia',
            'Greenock Creek',
            'Oliverhill',
            'Tait',
            'Wildberry Estate',
            'NEWBLOOD',
            'Rieslingfreak'
        ]
        # Also store shortened versions for matching
        self.producer_variants = {
            'BLACK PEARL': 'Black Pearl Vineyards',
            'DOWNES FAMILY': 'Downes Family Vineyards',
            'CASAS DEL BOSQUE': 'Casas del Bosque',
            'WILDBERRY': 'Wildberry Estate',
            'VINA ALICIA': 'Vina Alicia',
            'NUGAN': 'Nugan Estate',
            'GREENOCK': 'Greenock Creek',
            'NEW BLOOD': 'NEWBLOOD',
            'NEWBLOOD NON-ALCOHOLIC': 'NEWBLOOD',
            'ANTUCURA CHERIE': 'Antucura',
            'MT FISHTAIL': 'Mount Fishtail'
        }
        
        # Dictionary to store learned URL patterns
        self.url_patterns = {}
        self.learned_urls = {}
        
    def find_producer_in_text(self, text):
        """Find the producer in the text using our known list"""
        # First check for exact matches
        for producer in self.producers:
            if producer.upper() in text:
                return producer
                
        # Then check for known variants
        for variant, full_name in self.producer_variants.items():
            if variant in text:
                return full_name
                
        return "UNKNOWN"
        
    def parse_description(self, description):
        """Parse the description into vintage, producer, and product components"""
        parts = description.split()
        
        # First part is usually the vintage (year or N/V)
        vintage = parts[0]
        if vintage.startswith('20') or vintage == 'N/V':
            parts = parts[1:]  # Remove vintage from parts
        else:
            vintage = ''
            
        # Find the producer using our list
        full_desc = ' '.join(parts)
        producer = self.find_producer_in_text(full_desc.upper())
        
        # Remove producer from description to get product name
        product = full_desc
        if producer != "UNKNOWN":
            # Remove the producer name (and any variants) from the product description
            product = product.replace(producer, '')
            for variant in self.producer_variants.keys():
                product = product.replace(variant, '')
        product = ' '.join(product.split())  # Clean up extra spaces
        
        return vintage, producer, product
        
    def extract_pdf_data(self, pdf_path):
        """Extract data from PDF using direct text extraction"""
        print(f"\nExtracting data from {pdf_path}...")
        data = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    lines = text.split('\n')
                    
                    for line in lines:
                        # Skip header lines and empty lines
                        if not line.strip() or 'DESCRIPTION' in line or 'PAGE' in line or 'RUN' in line:
                            continue
                            
                        # Each line should start with an 'S' followed by alphanumeric characters
                        if line[0] == 'S':
                            try:
                                # Split by spaces, but keep the description together
                                parts = line.split()
                                if len(parts) >= 4:  # Make sure we have at least SKU, description, and some numbers
                                    sku = parts[0]
                                    
                                    # The last three numbers are on_hand, on_order, available
                                    available = parts[-1]
                                    on_order = parts[-2]
                                    on_hand = parts[-3]
                                    
                                    # Everything in between is the description
                                    description = ' '.join(parts[1:-3])
                                    
                                    # Parse the description into components
                                    vintage, producer, product = self.parse_description(description)
                                    
                                    data.append({
                                        'SKU': sku,
                                        'Vintage': vintage,
                                        'Producer': producer,
                                        'Product': product,
                                        'Full Description': description,
                                        'On Hand': on_hand,
                                        'On Order': on_order,
                                        'Available': available,
                                        'On Website': False,
                                        'Has Spec Sheet': False,
                                        'Has Shelf-Talker': False,
                                        'Has Hi-Res Label': False,
                                        'Has Bottle Shot': False,
                                        'Product URL': '',
                                        'Varietal Mismatch': False
                                    })
                            except Exception as e:
                                print(f"Skipping line due to error: {str(e)}")
                                continue
                                
        except Exception as e:
            print(f"Error extracting PDF data: {str(e)}")
            return None
            
        if not data:
            print("\nNo data was extracted from the PDF!")
            return None
            
        print(f"\nExtraction complete. Found {len(data)} products.")
        return pd.DataFrame(data)

    def setup_selenium(self):
        """Initialize Selenium WebDriver with retry logic"""
        print("\nSetting up web browser...")
        for attempt in range(self.max_retries):
            try:
                if self.driver is not None:
                    try:
                        self.driver.quit()
                    except:
                        pass
                options = webdriver.ChromeOptions()
                options.add_argument('--headless')
                options.add_argument('--disable-gpu')
                options.add_argument('--no-sandbox')
                options.add_argument('--disable-dev-shm-usage')
                options.add_argument('--window-size=1920,1080')
                # Performance optimizations
                options.add_argument('--disable-extensions')
                options.add_argument('--disable-images')
                options.add_argument('--disable-javascript')
                options.add_argument('--blink-settings=imagesEnabled=false')
                options.page_load_strategy = 'eager'
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
                self.driver.set_page_load_timeout(self.page_load_timeout)
                self.wait = WebDriverWait(self.driver, 10)
                return True
            except Exception as e:
                print(f"Attempt {attempt + 1} failed: {str(e)}")
                time.sleep(2)
        return False

    def safe_get_url(self, url, max_retries=None):
        """Safely navigate to a URL with retry logic"""
        if max_retries is None:
            max_retries = self.max_retries
            
        for attempt in range(max_retries):
            try:
                if not self.driver or not self.is_session_valid():
                    if not self.setup_selenium():
                        return False
                print(f"\nLoading {url}...")
                self.driver.get(url)
                self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                return True
            except Exception as e:
                print(f"Attempt {attempt + 1} failed to load {url}: {str(e)}")
                time.sleep(2)
                if "timeout" in str(e).lower():
                    print("Page load timed out, retrying with longer timeout...")
                    try:
                        self.driver.set_page_load_timeout(self.page_load_timeout * 1.5)
                    except:
                        pass
        return False

    def is_session_valid(self):
        """Check if the current session is valid"""
        try:
            # Try to access a simple property
            _ = self.driver.current_url
            return True
        except:
            return False

    def get_page_content(self):
        """Get all text content from the page for debugging"""
        return self.driver.execute_script("""
            return document.body.innerText;
        """)
        
    def get_producer_page_url(self, producer):
        """Convert producer name to URL format"""
        url_mapping = {
            'Mount Fishtail': 'mount-fishtail-wines',
            'Sherwood': 'sherwood-wines',
            'Stratum': 'stratum-wines',
            'Black Pearl Vineyards': 'black-pearl-wines',
            'Downes Family Vineyards': 'downes-family-vineyards-wines',
            'David Finlayson': 'david-finlayson-wines',
            'Luddite': 'luddite-wines',
            'Painted Wolf': 'painted-wolf-wines',
            'Thistledown': 'thistledown-wines',
            'Nugan Estate': 'nugan-wines',
            'RL Buller': 'rl-buller-wines',
            'Water Wheel': 'water-wheel-wines',
            'Casas del Bosque': 'casas-del-bosque-wines',
            'Earthsong': 'earthsong-wines',
            'Miha': 'miha-wines-2',
            'Almarada': 'almarada-wines',
            'Antucura': 'antucura-wines',
            'Vina Alicia': 'vina-alicia-2',
            'Greenock Creek': 'greenock-wines',
            'Oliverhill': 'oliver-hill-wines',
            'Tait': 'tait-wines',
            'Wildberry Estate': 'wildberry-estate-wines',
            'NEWBLOOD': 'newblood-wines',
            'Rieslingfreak': 'rieslingfreak-wines'
        }
        
        if producer in url_mapping:
            return f"{self.base_url}/{url_mapping[producer]}/"
        return None

    def get_producer_products(self, producer):
        """Get all products for a producer from the website with improved error handling"""
        producer_url = self.get_producer_page_url(producer)
        if not producer_url:
            print(f"No URL mapping found for producer: {producer}")
            return []
            
        try:
            print(f"\nChecking {producer_url}...")
            if not self.safe_get_url(producer_url):
                return []
                
            print(f"Page title: {self.driver.title}")
            print(f"Current URL: {self.driver.current_url}")
            
            products = []
            
            # Wait for either wine elements or buttons to be present
            try:
                elements = self.wait.until(
                    EC.presence_of_all_elements_located((
                        By.CSS_SELECTOR, 
                        'article[type-wines], .elementor-button-wrapper a'
                    ))
                )
                
                # Process wine elements
                wine_elements = self.driver.find_elements(By.CSS_SELECTOR, 'article[type-wines]')
                if wine_elements:
                    for element in wine_elements:
                        try:
                            title = element.find_element(By.CSS_SELECTOR, '.elementor-heading-title').text
                            if not title:
                                continue
                                
                            view_btn = element.find_element(By.CSS_SELECTOR, '.elementor-button-wrapper a')
                            url = view_btn.get_attribute('href')
                            
                            print(f"Found wine: {title}")
                            products.append({
                                'name': title,
                                'url': url
                            })
                            
                        except Exception as e:
                            continue
                
                # If no products found, try buttons approach
                if not products:
                    print("Trying alternate approach with View Wine buttons...")
                    buttons = self.driver.find_elements(By.CSS_SELECTOR, '.elementor-button-wrapper a')
                    for button in buttons:
                        try:
                            container = button.find_element(By.XPATH, './ancestor::article')
                            title = container.find_element(By.CSS_SELECTOR, '.elementor-heading-title').text
                            url = button.get_attribute('href')
                            
                            if title and url:
                                print(f"Found wine via button: {title}")
                                products.append({
                                    'name': title,
                                    'url': url
                                })
                                
                        except Exception:
                            continue
                            
            except TimeoutException:
                print(f"Timeout waiting for products on {producer_url}")
                
            print(f"\nFinished processing {producer}. Found {len(products)} wines.")
            return products
            
        except Exception as e:
            print(f"Error getting products for {producer}: {e}")
            return []

    def check_product_details(self, product_url):
        """Check if a product has all required assets with retry logic"""
        results = {
            'Has Spec Sheet': False,
            'Has Shelf-Talker': False,
            'Has Hi-Res Label': False,
            'Has Bottle Shot': False
        }

        if not self.safe_get_url(product_url):
            return results

        try:
            # Wait for trade tools to be present
            try:
                trade_tools = self.wait.until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.elementor-button-wrapper a'))
                )
            except TimeoutException:
                return results

            # Instead of clicking, we'll check the href attributes directly
            for tool in trade_tools:
                try:
                    tool_text = tool.text.strip().lower()
                    href = tool.get_attribute('href')
                    
                    if not href:
                        continue
                        
                    tool_map = {
                        'spec sheet': 'Has Spec Sheet',
                        'shelf-talker': 'Has Shelf-Talker',
                        'hi-res label': 'Has Hi-Res Label',
                        'bottle shot': 'Has Bottle Shot'
                    }

                    for text, key in tool_map.items():
                        if text in tool_text:
                            # If we find a matching button with an href, consider it available
                            results[key] = True
                            break

                except Exception:
                    continue

            return results

        except Exception as e:
            print(f"Error checking product details at {product_url}: {e}")
            return results

    def set_progress_callback(self, callback_function):
        """Set a callback function to report progress
        Callback should accept (current, total, producer) parameters
        """
        self.progress_callback = callback_function
        
    def detect_varietal(self, text):
        """Detect the varietal from product name or URL"""
        text = text.lower()
        
        # Check for each varietal term in the text
        for term, varietal in self.varietal_terms.items():
            if term in text:
                return varietal
                
        return None
        
    def check_varietal_mismatch(self, product_name, url):
        """Check if there's a mismatch between product name varietal and URL varietal"""
        if not product_name or not url:
            return False
            
        product_varietal = self.detect_varietal(product_name)
        url_varietal = self.detect_varietal(url)
        
        if product_varietal and url_varietal and product_varietal != url_varietal:
            print(f"Varietal mismatch: {product_name} ({product_varietal}) vs {url} ({url_varietal})")
            return True
            
        return False

    def check_duplicate_urls(self):
        """Check if any URL is assigned to multiple SKUs that aren't allowed to be duplicates"""
        url_to_skus = {}
        duplicates = []
        
        # Find all non-empty URLs in mapping
        for sku, url in self.sku_url_mapping.items():
            if url and url != 'NO_MATCH':
                if url in url_to_skus:
                    url_to_skus[url].append(sku)
                else:
                    url_to_skus[url] = [sku]
        
        # Check for duplicates
        for url, skus in url_to_skus.items():
            if len(skus) > 1:
                # Check if all SKUs in this group are allowed to share URLs
                is_allowed = True
                for i, sku1 in enumerate(skus):
                    for sku2 in skus[i+1:]:
                        # Check if this pair of SKUs is allowed to share a URL
                        if (sku1 not in self.allowed_duplicate_skus or 
                            sku2 not in self.allowed_duplicate_skus[sku1]):
                            is_allowed = False
                            break
                    if not is_allowed:
                        break
                
                if not is_allowed:
                    duplicates.append((url, skus))
                
        return duplicates

    def extract_sku_components(self, sku):
        """Extract the base part and vintage part from a SKU"""
        # Most SKUs follow pattern like S1BLACB23 where last 2 digits are vintage
        if len(sku) >= 9 and sku[-2:].isdigit():
            return sku[:-2], sku[-2:]
            
        # Some SKUs might have vintage with a separator
        match = re.search(r'(.+?)[-_](\d{2})$', sku)
        if match:
            return match.group(1), match.group(2)
            
        return sku, None  # No clear vintage part
        
    def learn_url_patterns(self):
        """Learn URL patterns from existing SKU-URL mappings"""
        print("\nLearning URL patterns from existing mappings...")
        vintage_patterns = {}
        
        # Group SKUs by their base part
        for sku, url in self.sku_url_mapping.items():
            if url == 'NO_MATCH':
                continue
                
            base_sku, vintage = self.extract_sku_components(sku)
            if not vintage:
                continue
                
            # Store this mapping
            if base_sku not in vintage_patterns:
                vintage_patterns[base_sku] = []
                
            vintage_patterns[base_sku].append((vintage, url))
        
        # Analyze patterns for each base SKU
        for base_sku, vintage_urls in vintage_patterns.items():
            if len(vintage_urls) < 2:
                continue  # Need at least 2 examples to establish a pattern
                
            # Sort by vintage
            vintage_urls.sort(key=lambda x: x[0])
            
            # Check if URLs follow a consistent pattern
            pattern_found = False
            for i in range(len(vintage_urls) - 1):
                vintage1, url1 = vintage_urls[i]
                vintage2, url2 = vintage_urls[i + 1]
                
                # Try to find vintage in URL
                url1_parts = url1.split('-')
                url2_parts = url2.split('-')
                
                if len(url1_parts) != len(url2_parts):
                    continue
                
                # Check which part changes with vintage
                diff_index = None
                for j, (part1, part2) in enumerate(zip(url1_parts, url2_parts)):
                    if part1 != part2:
                        if diff_index is None:
                            diff_index = j
                        else:
                            # More than one part differs - not a simple pattern
                            diff_index = None
                            break
                
                if diff_index is not None:
                    # Check if the different part contains the vintage
                    if f"20{vintage1}" in url1_parts[diff_index] and f"20{vintage2}" in url2_parts[diff_index]:
                        # Found a pattern! Store it
                        pattern = url1_parts.copy()
                        pattern[diff_index] = f"YEAR_PLACEHOLDER"
                        pattern_str = '-'.join(pattern)
                        
                        self.url_patterns[base_sku] = {
                            'pattern': pattern_str,
                            'examples': vintage_urls,
                            'year_format': "20{vintage}",
                            'position': diff_index,
                            'vintages': [v for v, _ in vintage_urls]
                        }
                        
                        print(f"  Learned pattern for {base_sku}: {vintage1}→{vintage2} maps to {pattern_str}")
                        pattern_found = True
                        break
            
            if pattern_found:
                self.url_patterns[base_sku]['vintages'] = [v for v, _ in vintage_urls]
    
    def predict_url_from_pattern(self, sku):
        """Predict URL for a new vintage based on patterns from previous vintages"""
        base_sku, vintage = self.extract_sku_components(sku)
        
        # Check if we already predicted this URL
        if sku in self.learned_urls:
            return self.learned_urls[sku]
            
        if not vintage or base_sku not in self.url_patterns:
            return None
            
        # Get the pattern
        pattern_info = self.url_patterns[base_sku]
        
        # Check if we've already seen this vintage
        if vintage in pattern_info['vintages']:
            return None  # This should already be in our manual mappings
            
        # Create the predicted URL
        pattern = pattern_info['pattern']
        year_format = pattern_info['year_format'].format(vintage=vintage)
        
        url_parts = pattern.split('-')
        url_parts[pattern_info['position']] = year_format
        predicted_url = '-'.join(url_parts)
        
        # Add the base URL if needed
        if not predicted_url.startswith('http'):
            predicted_url = f"{self.base_url}/wines/{predicted_url}/"
            
        print(f"  Predicted URL for {sku}: {predicted_url}")
        
        # Verify the URL exists
        try:
            # First try a lightweight HEAD request
            response = requests.head(predicted_url, timeout=5)
            if response.status_code == 200:
                print(f"  Verified URL exists: {predicted_url}")
                self.learned_urls[sku] = predicted_url
                return predicted_url
        except:
            pass
            
        return None

    def process_inventory(self, pdf_path):
        """Process inventory from PDF and check against website"""
        print("\nProcessing inventory...")
        current_date = datetime.now().strftime('%m%d%y')
        self.current_date = current_date
        
        # Learn URL patterns from existing SKU-URL mappings
        self.learn_url_patterns()
        
        # Check for duplicate URL mappings before starting
        duplicates = self.check_duplicate_urls()
        if duplicates:
            print("\nWARNING: Found duplicate URL mappings:")
            for url, skus in duplicates:
                print(f"  URL: {url}")
                print(f"  SKUs: {', '.join(skus)}")
            print("\nPlease fix duplicate mappings before continuing.")
            return False
        
        # Extract data from PDF
        inventory_df = self.extract_pdf_data(pdf_path)
        if inventory_df is None or inventory_df.empty:
            print("Error: No data extracted from PDF!")
            return False
            
        # Setup webdriver
        if self.driver is None:
            self.setup_selenium()
        
        # Process by producer for more accurate results
        print("\nChecking products on website...")
        
        # Get all unique producers
        unique_producers = inventory_df['Producer'].unique()
        producers = [p for p in unique_producers if p != "UNKNOWN"]
        
        # Count of relevant products for progress tracking
        products_total = len(inventory_df[inventory_df['Producer'] != "UNKNOWN"])
        products_checked = 0
        
        # Create a dictionary to track all website products
        all_website_products = {}
        
        # Create a dictionary to track URL usage
        used_urls = set()
        
        # For each producer
        for producer in producers:
            print(f"\nProcessing {producer} products...")
            
            # Get all products for this producer
            producer_df = inventory_df[inventory_df['Producer'] == producer]
            
            # Get producer products from website
            website_products = self.get_producer_products(producer)
            
            # Store website products for later comparison
            all_website_products[producer] = website_products
            
            # Process products for this producer
            self.process_producer_products(inventory_df, producer_df, website_products, used_urls, products_checked, products_total)
            products_checked += len(producer_df)
        
        # Find website products not in inventory
        website_only_df = self.find_website_only_products(all_website_products, used_urls)
        
        # Generate Excel report
        return self.generate_excel_report(inventory_df, website_only_df)
                
    def process_producer_products(self, inventory_df, producer_df, website_products, used_urls, products_checked, products_total):
        """Process all products for a producer"""
        for index, row in producer_df.iterrows():
            sku = row['SKU']
            product_name = row['Product']
            
            # Skip specific SKU that should be excluded entirely
            if sku == 'S1EDGGCSM20':
                print(f"Skipping SKU {sku} as per configuration")
                continue
                
            # Try direct SKU mapping first
            if sku in self.sku_url_mapping:
                product_url = self.sku_url_mapping[sku]
                
                # Special case for products that should NOT be matched
                if product_url == 'NO_MATCH':
                    print(f"SKU {sku} intentionally excluded from website matching")
                    continue
                    
                self.process_mapped_product(inventory_df, index, sku, product_name, product_url, used_urls)
            else:
                # Try to predict URL from patterns
                predicted_url = self.predict_url_from_pattern(sku)
                if predicted_url:
                    print(f"Using pattern-predicted URL for {sku}: {predicted_url}")
                    self.process_mapped_product(inventory_df, index, sku, product_name, predicted_url, used_urls)
                else:
                    # Try to find a matching product on the website
                    self.find_matching_product(inventory_df, index, sku, row, product_name, website_products, used_urls)
            
            # Report progress through callback if available
            if self.progress_callback:
                self.progress_callback(products_checked + index - producer_df.index[0] + 1, products_total, producer_df.iloc[0]['Producer'])
    
    def process_mapped_product(self, inventory_df, index, sku, product_name, product_url, used_urls):
        """Process a product with a known URL mapping"""
        print(f"Using URL mapping for {sku}: {product_url}")
        used_urls.add(product_url)
        
        results = self.check_product_details(product_url)
        has_varietal_mismatch = self.check_varietal_mismatch(product_name, product_url)
        
        # Update the dataframe with results
        for key, value in results.items():
            inventory_df.at[index, key] = value
        
        inventory_df.at[index, 'On Website'] = True
        inventory_df.at[index, 'Product URL'] = product_url
        inventory_df.at[index, 'Varietal Mismatch'] = has_varietal_mismatch
    
    def find_matching_product(self, inventory_df, index, sku, row, product_name, website_products, used_urls):
        """Find a matching product on the website"""
        best_match = None
        best_match_score = 0
        
        for web_product in website_products:
            web_name = web_product['name'].lower()
            inv_name = product_name.lower()
            
            # Calculate simple matching score
            words_in_inv = set(inv_name.split())
            words_in_web = set(web_name.split())
            common_words = words_in_inv.intersection(words_in_web)
            score = len(common_words) / max(len(words_in_inv), len(words_in_web))
            
            # Consider vintage if available
            if row['Vintage'] and row['Vintage'] != 'N/V':
                if row['Vintage'] in web_product['name']:
                    score += 0.3  # Boost score if vintage matches
            
            if score > best_match_score:
                best_match_score = score
                best_match = web_product
        
        # If we found a reasonable match
        if best_match and best_match_score > 0.3:
            print(f"Match found for {sku} - {product_name}: {best_match['name']} (Score: {best_match_score:.2f})")
            used_urls.add(best_match['url'])
            
            # Check detailed info for this product
            results = self.check_product_details(best_match['url'])
            has_varietal_mismatch = self.check_varietal_mismatch(product_name, best_match['url'])
            
            # Update the dataframe with results
            for key, value in results.items():
                inventory_df.at[index, key] = value
            
            inventory_df.at[index, 'On Website'] = True
            inventory_df.at[index, 'Product URL'] = best_match['url']
            inventory_df.at[index, 'Varietal Mismatch'] = has_varietal_mismatch
        else:
            print(f"No match found for {sku} - {product_name}")
            inventory_df.at[index, 'Varietal Mismatch'] = False
    
    def find_website_only_products(self, all_website_products, used_urls):
        """Find products that are on the website but not in inventory"""
        website_only_products = []
        
        for producer, products in all_website_products.items():
            for product in products:
                if product['url'] not in used_urls and product['url'] not in self.sku_url_mapping.values():
                    website_only_products.append({
                        'Producer': producer,
                        'Product': product['name'],
                        'Product URL': product['url']
                    })
        
        # Create a DataFrame for website-only products
        if website_only_products:
            return pd.DataFrame(website_only_products)
        else:
            return pd.DataFrame(columns=['Producer', 'Product', 'Product URL'])
    
    def generate_excel_report(self, inventory_df, website_only_df):
        """Generate Excel report with all results"""
        print("\nGenerating Excel report...")
        try:
            # Sort the inventory DataFrame
            inventory_df = inventory_df.sort_values(['Producer', 'SKU'])
            
            # Save results to Excel with multiple sheets
            output_file = f'inventory_report_{self.current_date}.xlsx'
            i = 1
            while True:
                try:
                    current_file = output_file if i == 1 else f"{os.path.splitext(output_file)[0]}_{i}{os.path.splitext(output_file)[1]}"
                        
                    with pd.ExcelWriter(current_file, engine='openpyxl') as writer:
                        inventory_df.to_excel(writer, sheet_name='Inventory Report', index=False)
                        website_only_df.to_excel(writer, sheet_name='Website Only Products', index=False)
                        
                        # Format Inventory Report sheet
                        inventory_sheet = writer.sheets['Inventory Report']
                        self.apply_conditional_formatting(inventory_sheet, inventory_df)
                        
                    print(f"\nResults saved to {current_file}")
                    print(f"- Inventory Report: {len(inventory_df)} products")
                    print(f"- Website Only Products: {len(website_only_df)} products")
                    break
                except PermissionError:
                    i += 1
                    if i > 10:
                        raise Exception("Could not save file after 10 attempts")
                    continue
            
            return True
            
        except Exception as e:
            print(f"Error generating report: {e}")
            return False
            
        finally:
            if self.driver:
                self.driver.quit()
                
    def apply_conditional_formatting(self, worksheet, df):
        """Apply conditional formatting to highlight issues"""
        # Define fill colors
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        
        # Get column indices
        on_website_col = df.columns.get_loc('On Website') + 1  # +1 because Excel is 1-indexed
        varietal_mismatch_col = df.columns.get_loc('Varietal Mismatch') + 1
        
        # Apply conditional formatting
        for row in range(2, len(df) + 2):  # +2 because Excel is 1-indexed and has a header row
            # Products not on website (missing)
            if not df.iloc[row-2]['On Website']:
                worksheet.cell(row=row, column=on_website_col).fill = red_fill
                
            # Products with varietal mismatch
            if df.iloc[row-2].get('Varietal Mismatch', False):
                worksheet.cell(row=row, column=varietal_mismatch_col).fill = yellow_fill

def main():
    parser = argparse.ArgumentParser(description='Process inventory PDF and check website')
    parser.add_argument('pdf_path', help='Path to the inventory PDF file')
    args = parser.parse_args()
    
    if not os.path.exists(args.pdf_path):
        print(f"Error: File not found: {args.pdf_path}")
        return
        
    checker = InventoryChecker()
    checker.process_inventory(args.pdf_path)

if __name__ == "__main__":
    main() 

# python inventory_checker.py "C:\Users\warwi\OneDrive\Desktop\itswarwick\Starz Updater\D916818--A.pdf"
